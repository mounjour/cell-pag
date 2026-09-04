from io import BytesIO
from urllib.parse import urlencode

from django.http import HttpResponse
from django.utils import timezone
from django.views import View
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.usuarios.mixins import DonoRequeridoMixin

from .forms import PeriodoForm
from .servicos import montar_relatorio


def _contexto(request):
    dados = request.GET or {
        "periodo": "diario",
        "referencia": timezone.localdate().isoformat(),
    }
    form = PeriodoForm(dados)
    if not form.is_valid():
        return {"form": form, "relatorio": None, "querystring": ""}
    relatorio = montar_relatorio(form.cleaned_data["inicio"], form.cleaned_data["fim"])
    querystring = urlencode(
        {
            "periodo": form.cleaned_data["periodo"],
            "referencia": form.cleaned_data["referencia"].isoformat(),
            "inicio": form.cleaned_data["inicio"].isoformat(),
            "fim": form.cleaned_data["fim"].isoformat(),
        }
    )
    return {"form": form, "relatorio": relatorio, "querystring": querystring}


class RelatorioView(DonoRequeridoMixin, TemplateView):
    template_name = "relatorios/painel.html"

    def get_context_data(self, **kwargs):
        return {**super().get_context_data(**kwargs), **_contexto(self.request)}


class RelatorioExcelView(DonoRequeridoMixin, View):
    def get(self, request):
        ctx = _contexto(request)
        if ctx["relatorio"] is None:
            return HttpResponse("Período inválido.", status=400)
        rel = ctx["relatorio"]
        wb = Workbook()
        resumo = wb.active
        resumo.title = "Resumo"
        resumo.append(["Relatório de pagamentos"])
        resumo.append(["Período", rel["inicio"], rel["fim"]])
        resumo.append([])
        resumo.append(["Indicador", "Valor"])
        indicadores = [
            ("Total previsto", rel["total_previsto"]),
            ("Total recebido", rel["total_recebido"]),
            ("Total em atraso", rel["total_atrasado"]),
            ("Pagamentos registrados", rel["quantidade_recebimentos"]),
            ("Parcelas em atraso", rel["quantidade_atrasados"]),
            ("Novos clientes", rel["novos_clientes"]),
            ("Contratos quitados", rel["contratos_quitados"]),
        ]
        for item in indicadores:
            resumo.append(item)

        recebidos = wb.create_sheet("Recebimentos")
        recebidos.append(["Data", "Cliente", "Contrato", "Parcela", "Forma", "Valor"])
        for pagamento in rel["recebimentos"]:
            recebidos.append([
                pagamento.data_pagamento,
                pagamento.contrato.cliente.nome,
                pagamento.contrato.apelido,
                pagamento.vencimento.numero if pagamento.vencimento else "",
                pagamento.get_forma_display(),
                pagamento.valor_pago,
            ])

        atrasados = wb.create_sheet("Em atraso")
        atrasados.append(
            ["Vencimento", "Cliente", "Contrato", "Parcela", "Previsto", "Pago", "Em aberto"]
        )
        for vencimento in rel["atrasados"]:
            atrasados.append([
                vencimento.data_vencimento,
                vencimento.contrato.cliente.nome,
                vencimento.contrato.apelido,
                vencimento.numero,
                vencimento.valor_previsto,
                vencimento.valor_pago,
                vencimento.valor_em_aberto,
            ])

        for planilha in wb.worksheets:
            planilha.freeze_panes = "A2"
            for cell in planilha[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1E6E5A")
                cell.alignment = Alignment(vertical="center")
            for coluna in planilha.columns:
                largura = min(max(len(str(c.value or "")) for c in coluna) + 2, 40)
                planilha.column_dimensions[get_column_letter(coluna[0].column)].width = largura

        for row in resumo.iter_rows(min_row=5, max_row=7, min_col=2, max_col=2):
            row[0].number_format = 'R$ #,##0.00'
        for planilha in (recebidos, atrasados):
            for row in planilha.iter_rows(min_row=2):
                row[0].number_format = "dd/mm/yyyy"
        for row in recebidos.iter_rows(min_row=2, min_col=6, max_col=6):
            row[0].number_format = 'R$ #,##0.00'
        for row in atrasados.iter_rows(min_row=2, min_col=5, max_col=7):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'

        arquivo = BytesIO()
        wb.save(arquivo)
        nome = f"relatorio-{rel['inicio']:%Y-%m-%d}-a-{rel['fim']:%Y-%m-%d}.xlsx"
        resposta = HttpResponse(
            arquivo.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resposta["Content-Disposition"] = f'attachment; filename="{nome}"'
        return resposta


class RelatorioPDFView(DonoRequeridoMixin, View):
    def get(self, request):
        ctx = _contexto(request)
        if ctx["relatorio"] is None:
            return HttpResponse("Período inválido.", status=400)
        rel = ctx["relatorio"]
        arquivo = BytesIO()
        _montar_pdf(arquivo, rel)
        nome = f"relatorio-{rel['inicio']:%Y-%m-%d}-a-{rel['fim']:%Y-%m-%d}.pdf"
        resposta = HttpResponse(arquivo.getvalue(), content_type="application/pdf")
        resposta["Content-Disposition"] = f'attachment; filename="{nome}"'
        return resposta


def _montar_pdf(arquivo, relatorio):
    estilos = getSampleStyleSheet()
    estilos.add(ParagraphStyle(name="Direita", parent=estilos["BodyText"], alignment=TA_RIGHT))
    documento = SimpleDocTemplate(
        arquivo,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=14 * mm,
        title="Relatório de pagamentos",
    )
    itens = [
        Paragraph("Relatório de pagamentos", estilos["Title"]),
        Paragraph(
            f"Período de {relatorio['inicio']:%d/%m/%Y} a {relatorio['fim']:%d/%m/%Y}",
            estilos["BodyText"],
        ),
        Spacer(1, 5 * mm),
    ]
    resumo = [
        ["Indicador", "Valor", "Indicador", "Valor"],
        ["Total previsto", _reais(relatorio["total_previsto"]), "Total recebido", _reais(relatorio["total_recebido"])],
        ["Total em atraso", _reais(relatorio["total_atrasado"]), "Parcelas atrasadas", str(relatorio["quantidade_atrasados"])],
        ["Novos clientes", str(relatorio["novos_clientes"]), "Contratos quitados", str(relatorio["contratos_quitados"])],
    ]
    itens.extend([_tabela_pdf(resumo, [60 * mm, 35 * mm, 60 * mm, 35 * mm]), Spacer(1, 6 * mm)])

    itens.append(Paragraph("Recebimentos", estilos["Heading2"]))
    recebimentos = [["Data", "Cliente", "Contrato", "Forma", "Valor"]]
    recebimentos.extend(
        [
            p.data_pagamento.strftime("%d/%m/%Y"),
            p.contrato.cliente.nome,
            p.contrato.apelido,
            p.get_forma_display(),
            _reais(p.valor_pago),
        ]
        for p in relatorio["recebimentos"]
    )
    if len(recebimentos) == 1:
        recebimentos.append(["Nenhum recebimento no período.", "", "", "", ""])
    itens.extend([_tabela_pdf(recebimentos, [28 * mm, 68 * mm, 58 * mm, 35 * mm, 32 * mm]), Spacer(1, 6 * mm)])

    itens.append(Paragraph("Parcelas em atraso", estilos["Heading2"]))
    atrasados = [["Vencimento", "Cliente", "Contrato", "Parcela", "Em aberto"]]
    atrasados.extend(
        [
            v.data_vencimento.strftime("%d/%m/%Y"),
            v.contrato.cliente.nome,
            v.contrato.apelido,
            str(v.numero),
            _reais(v.valor_em_aberto),
        ]
        for v in relatorio["atrasados"]
    )
    if len(atrasados) == 1:
        atrasados.append(["Nenhuma parcela em atraso.", "", "", "", ""])
    itens.append(_tabela_pdf(atrasados, [28 * mm, 68 * mm, 58 * mm, 35 * mm, 32 * mm]))
    documento.build(itens, onFirstPage=_rodape_pdf, onLaterPages=_rodape_pdf)


def _tabela_pdf(dados, larguras):
    tabela = Table(dados, colWidths=larguras, repeatRows=1)
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E6E5A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D6D4C8")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1EFE6")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return tabela


def _rodape_pdf(canvas, documento):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#666666"))
    canvas.drawRightString(landscape(A4)[0] - 12 * mm, 7 * mm, f"Página {documento.page}")
    canvas.restoreState()


def _reais(valor):
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
