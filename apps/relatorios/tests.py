import datetime
from decimal import Decimal
from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook
from validate_docbr import CPF as CPFGen

from apps.clientes.models import Cliente
from apps.contratos.models import Contrato
from apps.pagamentos.models import Pagamento, Vencimento

from .forms import PeriodoForm
from .servicos import montar_relatorio


date = datetime.date


@pytest.fixture
def dono(django_user_model):
    return django_user_model.objects.create_user(
        "dono", password="s3nha-forte-123", perfil=django_user_model.Perfil.DONO
    )


@pytest.fixture
def dono_client(client, dono):
    client.force_login(dono)
    return client


@pytest.fixture
def dados_relatorio(db, dono):
    cliente = Cliente.objects.create(
        nome="Cliente Relatório",
        cpf=CPFGen().generate(),
        telefone_whatsapp="+5583999991000",
    )
    contrato = Contrato.objects.create(
        cliente=cliente,
        apelido="Celular teste",
        aparelho_modelo="Modelo teste",
        valor_total=Decimal("200.00"),
        estrutura=Contrato.Estrutura.DIARIA,
        valor_parcela=Decimal("100.00"),
        num_parcelas=2,
        data_inicio=date(2026, 8, 31),
    )
    vencido = Vencimento.objects.create(
        contrato=contrato,
        numero=1,
        data_vencimento=date(2026, 9, 1),
        valor_previsto=Decimal("100.00"),
    )
    pago = Vencimento.objects.create(
        contrato=contrato,
        numero=2,
        data_vencimento=date(2026, 9, 2),
        valor_previsto=Decimal("100.00"),
    )
    Pagamento(
        contrato=contrato,
        vencimento=pago,
        data_pagamento=date(2026, 9, 2),
        valor_pago=Decimal("100.00"),
        usuario_baixa=dono,
    ).registrar()
    return contrato, vencido, pago


def test_periodo_semanal_e_mensal():
    semanal = PeriodoForm({"periodo": "semanal", "referencia": "2026-09-04"})
    assert semanal.is_valid()
    assert semanal.cleaned_data["inicio"] == date(2026, 8, 31)
    assert semanal.cleaned_data["fim"] == date(2026, 9, 6)

    mensal = PeriodoForm({"periodo": "mensal", "referencia": "2026-02-10"})
    assert mensal.is_valid()
    assert mensal.cleaned_data["inicio"] == date(2026, 2, 1)
    assert mensal.cleaned_data["fim"] == date(2026, 2, 28)


@pytest.mark.django_db
def test_consolidacao_do_periodo(dados_relatorio):
    rel = montar_relatorio(date(2026, 9, 1), date(2026, 9, 2))
    assert rel["total_previsto"] == Decimal("200.00")
    assert rel["total_recebido"] == Decimal("100.00")
    assert rel["total_atrasado"] == Decimal("100.00")
    assert rel["quantidade_atrasados"] == 1


@pytest.mark.django_db
def test_relatorio_respeita_janela_semanal(dados_relatorio):
    contrato, vencimento, _ = dados_relatorio
    contrato.estrutura = Contrato.Estrutura.SEMANAL
    contrato.save(update_fields=["estrutura", "atualizado_em"])
    vencimento.data_vencimento = date(2026, 9, 1)
    vencimento.save(update_fields=["data_vencimento", "atualizado_em"])

    durante_a_semana = montar_relatorio(date(2026, 9, 1), date(2026, 9, 4))
    assert durante_a_semana["quantidade_atrasados"] == 0

    depois_do_domingo = montar_relatorio(date(2026, 9, 1), date(2026, 9, 7))
    assert depois_do_domingo["quantidade_atrasados"] == 1


@pytest.mark.django_db
def test_relatorio_historico_inclui_pagamento_feito_depois(dados_relatorio, dono):
    contrato, vencimento, _ = dados_relatorio
    Pagamento(
        contrato=contrato,
        vencimento=vencimento,
        data_pagamento=date(2026, 9, 5),
        valor_pago=Decimal("100.00"),
        usuario_baixa=dono,
    ).registrar()

    no_dia_anterior = montar_relatorio(date(2026, 9, 1), date(2026, 9, 4))
    assert no_dia_anterior["total_atrasado"] == Decimal("100.00")

    depois_do_pagamento = montar_relatorio(date(2026, 9, 1), date(2026, 9, 6))
    assert depois_do_pagamento["total_atrasado"] == Decimal("0.00")


@pytest.mark.django_db
def test_relatorio_exige_perfil_dono(auth_client):
    resposta = auth_client.get(reverse("relatorios:painel"))
    assert resposta.status_code == 403


@pytest.mark.django_db
def test_painel_do_dono(dono_client, dados_relatorio):
    resposta = dono_client.get(
        reverse("relatorios:painel"),
        {"periodo": "personalizado", "inicio": "2026-09-01", "fim": "2026-09-02"},
    )
    assert resposta.status_code == 200
    assert resposta.context["relatorio"]["total_recebido"] == Decimal("100.00")
    assert "Cliente Relatório" in resposta.content.decode()


@pytest.mark.django_db
def test_exportacao_excel(dono_client, dados_relatorio):
    resposta = dono_client.get(
        reverse("relatorios:excel"),
        {"periodo": "personalizado", "inicio": "2026-09-01", "fim": "2026-09-02"},
    )
    assert resposta.status_code == 200
    workbook = load_workbook(BytesIO(resposta.content), data_only=False)
    assert workbook.sheetnames == ["Resumo", "Recebimentos", "Em atraso"]
    assert workbook["Resumo"]["B5"].value == 200


@pytest.mark.django_db
def test_exportacao_pdf(dono_client, dados_relatorio):
    resposta = dono_client.get(
        reverse("relatorios:pdf"),
        {"periodo": "personalizado", "inicio": "2026-09-01", "fim": "2026-09-02"},
    )
    assert resposta.status_code == 200
    assert resposta.content.startswith(b"%PDF")


@pytest.mark.django_db
def test_quitacao_grava_data_real(dono_client, dados_relatorio, monkeypatch):
    contrato, _, _ = dados_relatorio
    monkeypatch.setattr("apps.contratos.views.timezone.localdate", lambda: date(2026, 9, 4))
    resposta = dono_client.post(reverse("contratos:quitar", args=[contrato.pk]))
    assert resposta.status_code == 302
    contrato.refresh_from_db()
    assert contrato.quitado_em == date(2026, 9, 4)
